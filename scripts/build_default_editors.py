#!/usr/bin/env python3
"""Build src-tauri/src/data/default_editors.json from spreadsheets or the local app database."""

from __future__ import annotations

import json
import os
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src-tauri/src/data/default_editors.json"

MAILBOX = Path("/Users/to/Downloads/投稿邮箱.xlsx")
DIRECTION = Path("/Users/to/Downloads/短篇小说收稿方向汇总_2026-08-13版.xlsx")
LIBRARY = Path("/Users/to/Downloads/编辑库.xlsx")

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
QQ_MAIL_RE = re.compile(r"(?:邮箱|email|mail)[:：\s]*(\d{5,12})", re.I)

# Exact aliases. Do not merge 掌心故事会/掌心雷, or 风行/风华.
PLATFORM_ALIASES = {
    "九州（一组）": "九州",
    "九州（二组）": "九州",
    "九州（海外）": "九州",
    "九州(一组)": "九州",
    "九州(二组)": "九州",
    "九州(海外)": "九州",
    "麦芽5组": "麦芽",
    "吾里鹿糖": "吾里",
    "长樂": "长乐",
    "花不完(刚刚好)": "花不完",
    "花不完（刚刚好）": "花不完",
    "GoodNovel(海外）": "GoodNovel",
    "GoodNovel(海外)": "GoodNovel",
    "GoodNovel（海外）": "GoodNovel",
    "GoodNovel（海外)": "GoodNovel",
    "dreame（海外）": "Dreame",
    "dreame(海外)": "Dreame",
    "dreame（海外)": "Dreame",
    "Dreame（海外）": "Dreame",
    "月下": "月下小说",
    "四季文学": "四季",
    "绣球阅读": "绣球",
    "17K": "17k",
    "长乐文学": "长乐",
    "长樂文学": "长乐",
    "花不完故事会": "花不完",
    "吾里文化": "吾里",
    "鹿糖": "吾里",
}

# Longer phrases first. Values are the tags written into work_type.
PHRASE_ALIASES: list[tuple[str, list[str]]] = [
    ("女频全类型", ["全品类"]),
    ("主要品类都收", ["全品类"]),
    ("bg全品类", ["全品类"]),
    ("不限题材", ["全品类"]),
    ("各类型", ["全品类"]),
    ("全类型", ["全品类"]),
    ("全品类", ["全品类"]),
    ("会员背叛", ["全员背叛"]),
    ("全家后悔", ["全员背叛"]),
    ("全员背叛", ["全员背叛"]),
    ("亲情对比虐", ["亲情虐"]),
    ("亲情虐", ["亲情虐"]),
    ("强情绪流", ["情绪流"]),
    ("高情绪", ["情绪流"]),
    ("强情绪", ["情绪流"]),
    ("情绪流", ["情绪流"]),
    ("虐爽", ["虐恋", "爽文"]),
]

# Do not include generic 故事/小说 (e.g. 「故事有看头」).
TAG_VOCAB = [
    "女频", "男频",
    "亲情虐", "情绪流", "大女主", "信息差", "区别对待", "全员背叛",
    "全品类", "死人文学",
    "追妻", "追夫", "世情", "爽文", "打脸", "虐恋", "甜宠", "婚恋",
    "古言", "现言", "年代", "种田", "宫斗", "宅斗", "萌宝", "总裁",
    "脑洞", "悬疑", "仙侠", "玄幻", "重生", "穿越", "穿书", "快穿",
    "校园", "都市", "科幻", "恐怖", "民间", "社会", "性转", "权谋",
    "民国", "修仙", "末世", "系统", "言情", "女强",
    "散文", "童话", "诗歌", "随笔", "纪实",
]


def cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def clean_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def extract_emails(text: str) -> list[str]:
    found = []
    for raw in EMAIL_RE.findall(text or ""):
        email = raw.strip(".,;；，。)）]】>》").lower()
        email = re.sub(r"(小窗|微信|qq|QQ)$", "", email)
        if email.count("@") == 1 and len(email) >= 6:
            found.append(email)
    if not found:
        for num in QQ_MAIL_RE.findall(text or ""):
            found.append(f"{num}@qq.com")
    out, seen = [], set()
    for email in found:
        if email not in seen:
            seen.add(email)
            out.append(email)
    return out


def extract_tags(notes: str) -> list[str]:
    """Derive work types from notes only."""
    blob = notes or ""
    tags, seen = [], set()

    def add(tag: str) -> None:
        tag = tag.strip()
        if tag and tag not in seen:
            seen.add(tag)
            tags.append(tag)

    remaining = blob
    for phrase, mapped in PHRASE_ALIASES:
        if phrase in remaining:
            for tag in mapped:
                add(tag)
            remaining = remaining.replace(phrase, " ")

    for tag in TAG_VOCAB:
        if tag in remaining:
            add(tag)
    return tags


def join_notes(parts: list[tuple[str, str]]) -> str:
    lines = []
    for label, raw in parts:
        text = clean_space(raw.replace("\n", " "))
        if not text:
            continue
        if len(text) > 280:
            text = text[:277] + "…"
        lines.append(f"{label}：{text}" if label else text)
    notes = "\n".join(lines)
    return notes[:800]


def normalize_name(name: str) -> str:
    return re.sub(r"[\s\[\]【】（）()·・\-—_]", "", name).lower()


def normalize_platform_key(platform: str) -> str:
    return re.sub(r"[\s（）()]", "", platform).replace("中文网", "").replace("文学", "")


def canonicalize_platform(platform: str) -> str:
    value = clean_space(platform).replace("樂", "乐")
    if not value:
        return ""
    if value in PLATFORM_ALIASES:
        return PLATFORM_ALIASES[value]
    stripped = re.sub(r"[\(（][^)）]*[\)）]\s*$", "", value).strip()
    stripped = re.sub(r"\d+组$", "", stripped).strip() or stripped
    if stripped in PLATFORM_ALIASES:
        return PLATFORM_ALIASES[stripped]
    if stripped:
        value = stripped
    return PLATFORM_ALIASES.get(value, value)


def sheet_rows(path: Path, name: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[name]
    rows = [[cell(v) for v in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def add_editor(bucket: dict[str, dict], item: dict) -> None:
    email = item["email"].strip().lower()
    if not email or "@" not in email:
        return
    item["email"] = email
    item["platform"] = canonicalize_platform(item.get("platform", ""))
    item["name"] = clean_space(item.get("name", ""))
    existing = bucket.get(email)
    if not existing:
        bucket[email] = item
        return
    if not existing["platform"]:
        existing["platform"] = item["platform"]
    if not existing["name"]:
        existing["name"] = item["name"]
    notes = [existing.get("notes", ""), item.get("notes", "")]
    existing["notes"] = join_notes([("", n) for n in notes if n])


DROPPED_EXPORT_TAGS = {"小程序", "知乎风", "番茄风", "其他", "知乎文", "海外投稿"}
EXPORT_TAG_ALIASES: dict[str, list[str]] = {
    "虐爽": ["虐恋", "爽文"],
    "古言知乎": ["古言"],
    "古言脑洞": ["古言", "脑洞"],
    "古言重生爽文": ["古言", "重生", "爽文"],
    "古言大女主": ["古言", "大女主"],
    "现言-甜文": ["现言", "甜宠"],
    "双男主": ["双男主"],
}
PLACEHOLDER_LOCALS = {"投稿", "editor", "test", "admin", "noreply", "no-reply"}


def parse_export_tags(raw: str) -> list[str]:
    tags, seen = [], set()

    def add(tag: str) -> None:
        tag = tag.strip()
        if not tag or tag in seen or tag in DROPPED_EXPORT_TAGS:
            return
        seen.add(tag)
        tags.append(tag)

    for part in re.split(r"[、，,|/]", raw or ""):
        piece = clean_space(part)
        if piece in EXPORT_TAG_ALIASES:
            for tag in EXPORT_TAG_ALIASES[piece]:
                add(tag)
        else:
            add(piece)
    return tags


def parse_exported_library() -> list[dict]:
    """Read 编辑库.xlsx. Only keep rows that already have a platform."""
    rows = sheet_rows(LIBRARY, "Sheet1")
    out = []
    for row in rows[1:]:
        row += [""] * 5
        platform, name, email, _style, work = row[:5]
        platform = canonicalize_platform(platform)
        if not platform:
            continue
        emails = extract_emails(email)
        if not emails:
            continue
        work_type = parse_export_tags(work)
        if "短篇" not in work_type and "中短篇" not in work_type:
            work_type = ["短篇", "中短篇", *work_type]
        for addr in emails:
            local = addr.split("@", 1)[0]
            if local in PLACEHOLDER_LOCALS:
                continue
            out.append({
                "platform": platform,
                "name": clean_space(name),
                "email": addr,
                "work_type": work_type,
                "notes": "",
            })
    return out


def merge_work_types(current: list[str], extra: list[str]) -> list[str]:
    seen = set()
    out = []
    for tag in [*current, *extra]:
        tag = tag.strip()
        if not tag or tag in seen or tag in DROPPED_EXPORT_TAGS:
            continue
        seen.add(tag)
        out.append(tag)
    return out


BOTH_CHANNEL_RE = re.compile(r"不限男女频|男女频|男频[、,/]女频|女频[、,/]男频|男/女频|女/男频|男频女频都")
REJECT_MALE_RE = re.compile(r"不收男频|不要男频|拒收男频")
REJECT_FEMALE_RE = re.compile(r"不收女频|不要女频|拒收女频")


def infer_channel_tags(notes: str) -> list[str]:
    text = notes or ""
    tags: list[str] = []
    if BOTH_CHANNEL_RE.search(text):
        if not REJECT_FEMALE_RE.search(text):
            tags.append("女频")
        if not REJECT_MALE_RE.search(text):
            tags.append("男频")
    if "女频" in text and "女频" not in tags and not REJECT_FEMALE_RE.search(text):
        tags.append("女频")
    if "男频" in text and "男频" not in tags and not REJECT_MALE_RE.search(text):
        tags.append("男频")
    return tags


def channel_tags_from_library() -> dict[str, list[str]]:
    if not LIBRARY.exists():
        return {}
    rows = sheet_rows(LIBRARY, "Sheet1")
    out: dict[str, list[str]] = {}
    for row in rows[1:]:
        row += [""] * 5
        _platform, _name, email, _style, work = row[:5]
        tags = [tag for tag in ("女频", "男频") if tag in (work or "")]
        for addr in extract_emails(email):
            if tags:
                out[addr] = tags
    return out


def apply_channel_tags() -> None:
    """Write 男频/女频 into existing work_type. No new field."""
    if not OUT.exists():
        raise SystemExit(f"missing bundled library: {OUT}")
    editors = json.loads(OUT.read_text(encoding="utf-8"))
    from_xlsx = channel_tags_from_library()
    added = {"女频": 0, "男频": 0}
    for editor in editors:
        email = editor["email"].strip().lower()
        current = editor.get("work_type") or []
        extra = merge_work_types(
            from_xlsx.get(email) or [],
            infer_channel_tags(editor.get("notes") or ""),
        )
        if not extra:
            extra = ["女频"]
        next_types = merge_work_types(current, extra)
        for tag in extra:
            if tag not in current:
                added[tag] = added.get(tag, 0) + 1
        editor["work_type"] = next_types
    editors.sort(key=lambda e: (e.get("platform", ""), e.get("name", ""), e["email"]))
    OUT.write_text(json.dumps(editors, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    female = sum(1 for e in editors if "女频" in e.get("work_type", []))
    male = sum(1 for e in editors if "男频" in e.get("work_type", []))
    both = sum(1 for e in editors if "女频" in e.get("work_type", []) and "男频" in e.get("work_type", []))
    print(f"applied channel tags -> {OUT}")
    print(f"女频 {female}, 男频 {male}, both {both}, newly added {added}")


def merge_from_library() -> None:
    if not LIBRARY.exists():
        raise SystemExit(f"missing editor library spreadsheet: {LIBRARY}")
    if not OUT.exists():
        raise SystemExit(f"missing bundled library: {OUT}")

    existing = json.loads(OUT.read_text(encoding="utf-8"))
    bucket = {item["email"].strip().lower(): item for item in existing}
    added = 0
    tagged = 0
    for item in parse_exported_library():
        email = item["email"]
        current = bucket.get(email)
        if current:
            next_types = merge_work_types(current.get("work_type") or [], item["work_type"])
            if next_types != current.get("work_type"):
                current["work_type"] = next_types
                tagged += 1
            continue
        bucket[email] = item
        added += 1

    editors = sorted(bucket.values(), key=lambda e: (e.get("platform", ""), e.get("name", ""), e["email"]))
    OUT.write_text(json.dumps(editors, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"merged library -> {OUT}")
    print(f"total {len(editors)} (added {added}, work_type merged {tagged}, existing kept {len(existing)})")


def parse_short() -> list[dict]:
    rows = sheet_rows(MAILBOX, "短篇编辑")
    last_platform = ""
    out = []
    for row in rows[1:]:
        row += [""] * 7
        platform, name, contact, review, settlement, accept, bonus = row[:7]
        platform = canonicalize_platform(platform) or last_platform
        if platform:
            last_platform = platform
        emails = extract_emails(contact)
        if not emails:
            continue
        for email in emails:
            out.append({
                "platform": platform,
                "name": re.sub(r"(小窗|微信)$", "", clean_space(name)),
                "email": email,
                "notes": join_notes([
                    ("审稿", review),
                    ("结算", settlement),
                    ("收稿", accept),
                    ("全勤", bonus),
                ]),
            })
    return out


def parse_direction_index() -> list[dict]:
    rows = sheet_rows(DIRECTION, "编辑总表")
    out = []
    for row in rows[1:]:
        row += [""] * 10
        _seq, platform, name, contact, _date, core, tags, reject, tip, examples = row[:10]
        emails = extract_emails(contact)
        out.append({
            "platform": canonicalize_platform(platform),
            "name": clean_space(name),
            "emails": emails,
            "notes": join_notes([
                ("收稿", core),
                ("标签", tags),
                ("不收", reject),
                ("注意", tip),
                ("例文", examples),
            ]),
        })
    return out


def enrich_from_direction(bucket: dict[str, dict], extras: list[dict]) -> None:
    by_key = {}
    for editor in bucket.values():
        key = (normalize_platform_key(editor["platform"]), normalize_name(editor["name"]))
        by_key.setdefault(key, []).append(editor)

    for extra in extras:
        targets = []
        for email in extra["emails"]:
            if email in bucket:
                targets.append(bucket[email])
        if not targets:
            key = (normalize_platform_key(extra["platform"]), normalize_name(extra["name"]))
            targets = by_key.get(key, [])
        if not targets and extra["emails"]:
            for email in extra["emails"]:
                add_editor(bucket, {
                    "platform": extra["platform"],
                    "name": extra["name"],
                    "email": email,
                    "notes": extra["notes"],
                })
            continue
        for editor in targets:
            editor["notes"] = join_notes([("", editor.get("notes", "")), ("", extra["notes"])])


def main() -> None:
    if not MAILBOX.exists() or not DIRECTION.exists():
        raise SystemExit(f"missing source spreadsheet: {MAILBOX} / {DIRECTION}")

    bucket: dict[str, dict] = {}
    for item in parse_short():
        add_editor(bucket, item)
    enrich_from_direction(bucket, parse_direction_index())

    editors = []
    for item in bucket.values():
        notes = item.get("notes", "")
        editors.append({
            "platform": canonicalize_platform(item["platform"]),
            "name": item["name"],
            "email": item["email"],
            "work_type": extract_tags(notes),
            "notes": notes,
        })

    editors.sort(key=lambda e: (e["platform"], e["name"], e["email"]))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(editors, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    platforms = Counter(e["platform"] or "未填" for e in editors)
    tags = Counter(t for e in editors for t in e["work_type"])
    print(f"wrote {len(editors)} editors -> {OUT}")
    print("platforms", len(platforms))
    print("top work_type", tags.most_common(30))
    print("empty work_type", sum(1 for e in editors if not e["work_type"]))
    print("empty notes", sum(1 for e in editors if not e["notes"]))
    print("canonical samples", {
        name: platforms.get(name, 0)
        for name in ("九州", "麦芽", "吾里", "长乐", "花不完", "GoodNovel", "Dreame", "月下小说", "四季", "绣球")
    })


def local_db_path() -> Path:
    if override := os.environ.get("NOVELSUB_DB"):
        return Path(override).expanduser()
    return Path.home() / "Library/Application Support/com.novelsub.desktop/novelsub.sqlite"


def export_from_local_db() -> None:
    import sqlite3

    db = local_db_path()
    if not db.exists():
        raise SystemExit(f"local editor database not found: {db}")

    conn = sqlite3.connect(db)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT platform, name, email, work_type, notes
        FROM editors
        WHERE TRIM(COALESCE(email, '')) != ''
        ORDER BY platform, name, email
        """
    ).fetchall()

    editors = []
    seen: set[str] = set()
    for row in rows:
        email = (row["email"] or "").strip().lower()
        if not email or email in seen:
            continue
        seen.add(email)
        try:
            work_type = json.loads(row["work_type"] or "[]")
        except json.JSONDecodeError:
            work_type = []
        if not isinstance(work_type, list):
            work_type = []
        editors.append({
            "platform": (row["platform"] or "").strip(),
            "name": (row["name"] or "").strip(),
            "email": email,
            "work_type": [str(tag).strip() for tag in work_type if str(tag).strip()],
            "notes": (row["notes"] or "").strip(),
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(editors, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    platforms = Counter(e["platform"] or "未填" for e in editors)
    tags = Counter(t for e in editors for t in e["work_type"])
    print(f"wrote {len(editors)} editors from {db} -> {OUT}")
    print("platforms", len(platforms))
    print("top work_type", tags.most_common(20))


if __name__ == "__main__":
    import sys
    if "--merge-library" in sys.argv:
        merge_from_library()
    elif "--apply-channel" in sys.argv:
        apply_channel_tags()
    elif "--from-local-db" in sys.argv:
        export_from_local_db()
    else:
        main()
